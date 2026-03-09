"""
ThermoPilot AI — Excel export routes
Exports scenarios and audits as .xlsx files using openpyxl.
"""
from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import Response
from sqlalchemy.orm import Session, joinedload
from uuid import UUID
import io
from datetime import datetime

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from app.database import get_db
from app.core.security import get_current_user
from app.models.audit import Audit
from app.models.building import Building
from app.models.scenario import RenovationScenario
from app.models.user import User

router = APIRouter(prefix="/exports", tags=["exports"])

_HEADER_FILL = PatternFill("solid", fgColor="2B6CB0")
_HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
_ALT_FILL = PatternFill("solid", fgColor="EBF4FF")
_BORDER = Border(
    left=Side(style="thin", color="CBD5E0"),
    right=Side(style="thin", color="CBD5E0"),
    top=Side(style="thin", color="CBD5E0"),
    bottom=Side(style="thin", color="CBD5E0"),
)


def _style_header_row(ws, row: int, ncols: int):
    for col in range(1, ncols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = _BORDER


def _style_data_row(ws, row: int, ncols: int, alt: bool = False):
    for col in range(1, ncols + 1):
        cell = ws.cell(row=row, column=col)
        if alt:
            cell.fill = _ALT_FILL
        cell.alignment = Alignment(vertical="center")
        cell.border = _BORDER


def _autofit(ws):
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=0)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 40)


@router.get("/audits/xlsx")
def export_audits_xlsx(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Export all completed audits as an Excel file."""
    audits = (
        db.query(Audit)
        .options(joinedload(Audit.building))
        .filter(
            Audit.organization_id == current_user.organization_id,
            Audit.status == "completed",
        )
        .order_by(Audit.created_at.desc())
        .all()
    )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Audits"
    ws.row_dimensions[1].height = 22

    headers = [
        "Bâtiment", "Type", "Ville", "Surface (m²)", "Année constr.",
        "Classe DPE", "Classe GES",
        "Énergie primaire (kWhpe/m²/an)", "CO₂ (kgCO₂/m²/an)",
        "Conso. finale (kWh/an)", "Coût estimé (€/an)",
        "Date audit",
    ]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    _style_header_row(ws, 1, len(headers))

    for i, audit in enumerate(audits, 2):
        b = audit.building
        snap = audit.result_snapshot or {}

        row = [
            b.name if b else "—",
            b.building_type if b else "—",
            b.city if b else "—",
            float(b.heated_area_m2) if b and b.heated_area_m2 else None,
            b.construction_year if b else None,
            audit.computed_energy_label or "—",
            audit.computed_ghg_label or "—",
            snap.get("primary_energy_per_m2"),
            snap.get("co2_per_m2"),
            snap.get("total_final_kwh"),
            snap.get("estimated_annual_cost_eur"),
            audit.created_at.strftime("%d/%m/%Y") if audit.created_at else "—",
        ]
        for col, val in enumerate(row, 1):
            ws.cell(row=i, column=col, value=val)
        _style_data_row(ws, i, len(headers), alt=(i % 2 == 0))

    ws.freeze_panes = "A2"
    _autofit(ws)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return Response(
        content=buf.read(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f'attachment; filename="thermopilot_audits_{datetime.now().strftime("%Y%m%d")}.xlsx"'
        },
    )


@router.get("/scenarios/xlsx")
def export_scenarios_xlsx(
    audit_id: UUID = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Export renovation scenarios as an Excel comparison file."""
    q = db.query(RenovationScenario).filter(
        RenovationScenario.organization_id == current_user.organization_id
    )
    if audit_id:
        q = q.filter(RenovationScenario.audit_id == audit_id)
    scenarios = q.order_by(RenovationScenario.created_at.desc()).all()

    if not scenarios:
        raise HTTPException(status_code=404, detail="Aucun scénario à exporter")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Scénarios de rénovation"
    ws.row_dimensions[1].height = 22

    headers = [
        "Scénario", "Type", "Classe visée",
        "Coût total (€)", "Économies énergie (kWh/an)",
        "Économies annuelles (€/an)", "Retour sur invest. (ans)",
        "Réduction CO₂ (kgCO₂/an)", "Statut", "Date création",
    ]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    _style_header_row(ws, 1, len(headers))

    for i, s in enumerate(scenarios, 2):
        row = [
            s.name,
            s.scenario_type,
            s.target_energy_label or "—",
            float(s.estimated_total_cost_eur) if s.estimated_total_cost_eur else None,
            float(s.estimated_energy_savings_kwh) if s.estimated_energy_savings_kwh else None,
            float(s.estimated_annual_savings_eur) if s.estimated_annual_savings_eur else None,
            float(s.simple_payback_years) if s.simple_payback_years else None,
            float(s.estimated_co2_reduction_kg) if s.estimated_co2_reduction_kg else None,
            s.status,
            s.created_at.strftime("%d/%m/%Y") if s.created_at else "—",
        ]
        for col, val in enumerate(row, 1):
            ws.cell(row=i, column=col, value=val)
        _style_data_row(ws, i, len(headers), alt=(i % 2 == 0))

    ws.freeze_panes = "A2"
    _autofit(ws)

    # Add summary sheet
    ws2 = wb.create_sheet("Résumé")
    ws2["A1"] = "Rapport comparatif — ThermoPilot AI"
    ws2["A1"].font = Font(bold=True, size=14, color="1A365D")
    ws2["A2"] = f"Généré le {datetime.now().strftime('%d/%m/%Y à %H:%M')}"
    ws2["A2"].font = Font(italic=True, color="4A5568")
    ws2["A4"] = "Nombre de scénarios"
    ws2["B4"] = len(scenarios)
    if scenarios:
        total_cost = sum(float(s.estimated_total_cost_eur or 0) for s in scenarios)
        total_savings = sum(float(s.estimated_energy_savings_kwh or 0) for s in scenarios)
        best = min(
            (s for s in scenarios if s.simple_payback_years),
            key=lambda s: float(s.simple_payback_years),
            default=None,
        )
        ws2["A5"] = "Coût total (tous scénarios)"
        ws2["B5"] = total_cost
        ws2["A6"] = "Économies totales potentielles (kWh/an)"
        ws2["B6"] = total_savings
        if best:
            ws2["A7"] = "Meilleur retour sur investissement"
            ws2["B7"] = f"{best.name} ({best.simple_payback_years} ans)"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return Response(
        content=buf.read(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f'attachment; filename="thermopilot_scenarios_{datetime.now().strftime("%Y%m%d")}.xlsx"'
        },
    )
